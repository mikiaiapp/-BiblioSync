"""
Excel report generation module for BiblioSync.
Generates a structured spreadsheet summarizing copied books, errors, and metadata.
"""

from pathlib import Path
from typing import List, Tuple, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.database.models import Book
from src.utils.logger import logger

def generate_excel_report(
    output_path: Path,
    copied_books: List[Tuple[Book, str]],
    failed_books: List[Tuple[Book, str]],
    summary_data: Dict[str, Any]
) -> None:
    """
    Creates an Excel spreadsheet containing three sheets:
    - Sheet 1: Libros Copiados (Successfully copied books)
    - Sheet 2: Errores (Copy errors and failures)
    - Sheet 3: Resumen (Sync run statistics summary)
    
    Args:
        output_path (Path): Target path to save the Excel file.
        copied_books: List of tuples (Book, destination_path)
        failed_books: List of tuples (Book, error_message)
        summary_data: Dict containing execution summary metrics.
    """
    logger.info(f"Generating Excel report at {output_path}...")
    
    wb = openpyxl.Workbook()
    # Remove default sheet to build custom ones
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)
    
    # Styling definitions
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Blue
    error_header_fill = PatternFill(start_color="A61C1C", end_color="A61C1C", fill_type="solid") # Dark Red
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # 1. Sheet "Libros Copiados"
    ws_copied = wb.create_sheet(title="Libros Copiados")
    headers_copied = ["Nombre Original", "Carpeta de Origen", "Carpeta de Destino", "Nombre Final", "Tamaño"]
    ws_copied.append(headers_copied)
    
    for book, dest_path_str in copied_books:
        dest_path = Path(dest_path_str)
        ws_copied.append([
            book.file_name,
            str(Path(book.file_path).parent),
            str(dest_path.parent),
            dest_path.name,
            book.file_size
        ])
        
    # Style Sheet 1 Headers
    for col_num in range(1, len(headers_copied) + 1):
        cell = ws_copied.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # 2. Sheet "Errores"
    ws_errors = wb.create_sheet(title="Errores")
    headers_errors = ["Nombre del Archivo", "Carpeta de Origen", "Mensaje de Error"]
    ws_errors.append(headers_errors)
    
    for book, err_msg in failed_books:
        ws_errors.append([
            book.file_name,
            str(Path(book.file_path).parent),
            err_msg
        ])
        
    # Style Sheet 2 Headers
    for col_num in range(1, len(headers_errors) + 1):
        cell = ws_errors.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = error_header_fill
        cell.alignment = center_align

    # 3. Sheet "Resumen"
    ws_summary = wb.create_sheet(title="Resumen")
    headers_summary = ["Métrica", "Valor"]
    ws_summary.append(headers_summary)
    
    for key, val in summary_data.items():
        ws_summary.append([key, val])
        
    # Style Sheet 3 Headers
    for col_num in range(1, len(headers_summary) + 1):
        cell = ws_summary.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Apply general styles & auto-fit columns for all sheets
    for ws in [ws_copied, ws_errors, ws_summary]:
        ws.row_dimensions[1].height = 24  # Give header row extra breathing room
        
        # Style all data rows
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 20
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = Font(name=font_family, size=10)
                cell.border = thin_border
                
                # Special alignments and formatting
                if ws.title == "Libros Copiados":
                    if col_idx in (1, 4):  # Filenames
                        cell.alignment = left_align
                    elif col_idx in (2, 3):  # Paths
                        cell.alignment = left_align
                    elif col_idx == 5:  # File Size
                        cell.alignment = right_align
                        cell.number_format = '#,##0'  # Format as integer with thousands separator
                elif ws.title == "Errores":
                    cell.alignment = left_align
                elif ws.title == "Resumen":
                    if col_idx == 1:
                        cell.alignment = left_align
                        cell.font = Font(name=font_family, size=10, bold=True)
                    else:
                        cell.alignment = left_align
                        
        # Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = cell.value
                if val is not None:
                    # Format size display length in auto-fit calculation if it's formatted
                    if isinstance(val, int) and cell.number_format == '#,##0':
                        val_str = f"{val:,}"
                    else:
                        val_str = str(val)
                    if len(val_str) > max_len:
                        max_len = len(val_str)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    try:
        # Create parent directories if they don't exist
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info("Excel report saved successfully.")
    except Exception as e:
        logger.error(f"Failed to save Excel report: {e}")
        raise e
